"""survey-data-preprocessing canonical script.

Parses a raw survey export (native `03_execute__survey-response__*.xlsx`) that uses
the team's standard 3-row header convention:

- row 1: question code (e.g. `P1B2`), set only on the first column of a
  multi-column question block, blank on the rest of that block
- row 2: full question text, same forward-fill rule as row 1
- row 3: option label for multi-column blocks (e.g. checkbox item text), or a
  scale-range hint (e.g. `1 ~ 7`) for single-column scale questions
- row 4+: respondent data

Columns before the first question code (respondent id, timestamps, etc.) never
get a code even after forward-fill — that absence IS the meta-column marker.

This structure was reverse-engineered by reading the actual raw xlsx for
`26.GP.UXQ` (그룹 UX 품질 진단, survey2) and cross-checked against the
already-published `column_ledger.csv` from the 2026-08-17 validation round.
It has only been verified against that one project — see CHANGELOG.md and
`04_validation_notes.md` in the matching validation_runs folder before trusting
this on a project with a different header convention.

Usage:
    python build_cleaned_dataset.py <raw_xlsx> <out_dir> [--header-rows 3] [--sheet NAME]
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

import openpyxl

SCALE_HINT_RE = re.compile(r"^\s*\d+\s*~\s*\d+\s*$")


def load_rows(xlsx_path: Path, sheet: str | None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return ws, ws.max_row, ws.max_column


def read_header(ws, max_col: int, header_rows: int):
    """Forward-fill question code (row1) and question text (row2) across a
    multi-column block, but treat any column whose row2 introduces its OWN new
    text as a fresh block boundary — even if row1 has no code there.

    This matters because native survey exports append platform-level system
    columns (성별/연령대/중복 판단/답변 완성도, a trailing open-ended question)
    after the last "P"-coded question. Those columns never get a row1 code, so
    a naive forward-fill would wrongly attribute them to the previous question
    block. Requiring row2 to also be blank before treating a column as a
    continuation fixes this (verified against `26.GP.UXQ` survey2 raw data,
    where a naive fill inflated the last question's block from 1 column to 6).
    """

    def row_values(r):
        return [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]

    row1 = row_values(1)
    row2 = row_values(2)
    row3 = row_values(3) if header_rows >= 3 else [None] * max_col

    code_ff: list = []
    text_ff: list = []
    current_code = None
    current_text = None
    for c1, c2 in zip(row1, row2):
        is_new_block = c1 not in (None, "") or c2 not in (None, "")
        if is_new_block:
            current_code = c1 if c1 not in (None, "") else None
            current_text = c2 if c2 not in (None, "") else current_text
        code_ff.append(current_code)
        text_ff.append(current_text)
    return code_ff, text_ff, row3


def build_columns(code_ff, text_ff, option_row, max_col: int):
    """Return per-column metadata: code, text, option_label, is_meta, column_key.

    A single-choice/scale question with a "기타 답변" (other, write-in) companion
    column still has only ONE substantive response column — the write-in
    column shouldn't force the main column into `CODE__label` naming. Verified
    against `26.GP.UXQ` survey2 (e.g. `P3B2` main + `P3B2` 기타 답변 write-in
    both map to a single bare `P3B2` key for the main column in the validated
    2026-08-17 output).
    """
    code_counts: dict = {}
    for code, label in zip(code_ff, option_row):
        if code is not None and not _looks_like_writein_label(label):
            code_counts[code] = code_counts.get(code, 0) + 1

    columns = []
    for i in range(max_col):
        code = code_ff[i]
        text = text_ff[i]
        option_label = option_row[i]
        is_meta = code is None
        if is_meta:
            # meta columns carry their own name in row 2 (e.g. 응답자ID)
            column_key = text if text is not None else f"col_{i + 1}"
            question_code = column_key
            question_text = column_key
            option_label = None
        else:
            question_code = code
            question_text = text
            is_writein = _looks_like_writein_label(option_label)
            if is_writein or code_counts.get(code, 0) > 1:
                label = option_label if option_label not in (None, "") else f"item{i + 1}"
                column_key = f"{code}__{label}"
            else:
                column_key = code
        columns.append(
            {
                "index": i + 1,
                "column_key": column_key,
                "question_code": question_code,
                "question_text": question_text,
                "option_label": option_label if not is_meta else None,
                "is_meta": is_meta,
            }
        )
    return columns


REDACTION_RE = re.compile(r"^\s*\[[^\]]+\]\s*$")


def classify_question_type(code: str, cols_for_code: list, sample_rows: list, option_row) -> str:
    """Heuristic family-level classifier. Documented, not guaranteed universal —
    see module docstring. Verify against a project before trusting blindly."""
    idxs = [c["index"] - 1 for c in cols_for_code]

    # Operational fields (raffle/incentive PII collection, e.g. "경품 수령을 위한
    # 연락처 뒤 4자리") are not survey-question content — exclude them from
    # question-family typing. Detect via question text keywords first (works
    # regardless of redaction), and via PII-redaction placeholder tokens like
    # `[삭제]` as a fallback for cases with no obvious keyword.
    question_text = cols_for_code[0].get("question_text") or ""
    if any(kw in question_text for kw in ("경품", "번호 뒤", "전화번호", "연락처")):
        return "operational"

    all_vals = []
    for idx in idxs:
        all_vals.extend(row[idx] for row in sample_rows if row[idx] not in (None, ""))
    if all_vals:
        redacted = sum(1 for v in all_vals if isinstance(v, str) and REDACTION_RE.match(v))
        if redacted / len(all_vals) >= 0.5:
            return "operational"

    hint = option_row[idxs[0]] if idxs else None
    if len(idxs) == 1 and isinstance(hint, str) and SCALE_HINT_RE.match(hint):
        return "scale"

    if len(idxs) == 1:
        values = [row[idxs[0]] for row in sample_rows if row[idxs[0]] not in (None, "")]
        if not values:
            return "single-choice"
        numeric = sum(1 for v in values if isinstance(v, (int, float)))
        if numeric == len(values):
            return "scale"
        distinct = len(set(str(v) for v in values))
        if distinct >= max(1, int(0.5 * len(values))) and len(values) > 5:
            return "open-ended"
        return "single-choice"

    # multi-column block: checkbox-style (value repeats the option label / any
    # non-null marker) => multiple-choice; small-integer rank values => ranking.
    # Exclude free-text write-in columns (e.g. "기타 답변") from the check —
    # a single open write-in column inside an otherwise numeric ranking block
    # should not flip the whole family to multiple-choice.
    numeric_idxs = [i for i in idxs if not _looks_like_writein_label(option_row[i])]
    all_values = []
    for idx in numeric_idxs:
        all_values.extend(row[idx] for row in sample_rows if row[idx] not in (None, ""))
    if not all_values:
        return "multiple-choice"
    numeric_like = sum(1 for v in all_values if isinstance(v, (int, float)) and float(v) == int(v) and 0 < v <= 10)
    if numeric_like / len(all_values) >= 0.9:
        return "ranking"
    return "multiple-choice"


def _looks_like_writein_label(label) -> bool:
    if not isinstance(label, str):
        return False
    return "답변" in label or "기타 응답" in label or "직접" in label


def detect_group_candidates(columns, data_rows, max_candidates=6):
    """Segment/group candidates are demographic or platform-level system
    columns (성별, 연령대, 중복 판단, ...) — i.e. columns with no question code
    (`is_meta`) — not substantive survey questions that merely happen to have
    few response levels. Restricting to `is_meta` columns matches the
    convention observed in the 2026-08-17 validated output (성별/연령대/중복
    판단 only, never a Likert item like P3B3)."""
    candidates = []
    for col in columns:
        if not col["is_meta"]:
            continue
        idx = col["index"] - 1
        values = [row[idx] for row in data_rows if row[idx] not in (None, "")]
        if not values:
            continue
        levels = {}
        for v in values:
            key = str(v)
            levels[key] = levels.get(key, 0) + 1
        n_levels = len(levels)
        # No minimum fill-rate requirement: system QC fields like 중복 판단
        # are sparsely filled (mostly blank = normal case) but are still a
        # valid segment candidate in the original convention.
        if 2 <= n_levels <= 8:
            candidates.append(
                {
                    "column_key": col["column_key"],
                    "question_code": col["question_code"],
                    "question_text": col["question_text"],
                    "n_levels": n_levels,
                    "levels": dict(sorted(levels.items(), key=lambda kv: -kv[1])),
                }
            )
    candidates.sort(key=lambda c: c["n_levels"])
    return candidates[:max_candidates]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("raw_xlsx")
    parser.add_argument("out_dir")
    parser.add_argument("--header-rows", type=int, default=3)
    parser.add_argument("--sheet", default=None)
    args = parser.parse_args()

    raw_path = Path(args.raw_xlsx)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    ws, max_row, max_col = load_rows(raw_path, args.sheet)
    code_ff, text_ff, option_row = read_header(ws, max_col, args.header_rows)
    columns = build_columns(code_ff, text_ff, option_row, max_col)

    data_start_row = args.header_rows + 1
    data_rows = []
    for r in range(data_start_row, max_row + 1):
        data_rows.append([ws.cell(row=r, column=c).value for c in range(1, max_col + 1)])

    # cleaned_dataset.csv
    header = [c["column_key"] for c in columns]
    cleaned_path = out_dir / "cleaned_dataset.csv"
    with cleaned_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in data_rows:
            writer.writerow(["" if v is None else v for v in row])

    # column_ledger.csv
    ledger_path = out_dir / "column_ledger.csv"
    with ledger_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            ["column_index", "column_key", "question_code", "question_text", "option_label", "is_meta", "non_null_count", "unique_non_null"]
        )
        for col in columns:
            idx = col["index"] - 1
            vals = [row[idx] for row in data_rows]
            non_null = [v for v in vals if v not in (None, "")]
            writer.writerow(
                [
                    col["index"],
                    col["column_key"],
                    col["question_code"],
                    col["question_text"],
                    col["option_label"] or "",
                    col["is_meta"],
                    len(non_null),
                    len(set(str(v) for v in non_null)),
                ]
            )

    # question family classification
    code_to_cols: dict = {}
    for col in columns:
        if col["is_meta"]:
            continue
        code_to_cols.setdefault(col["question_code"], []).append(col)

    family_type_counts: dict = {}
    for code, cols_for_code in code_to_cols.items():
        qtype = classify_question_type(code, cols_for_code, data_rows, option_row)
        family_type_counts[qtype] = family_type_counts.get(qtype, 0) + 1

    operational_count = family_type_counts.pop("operational", 0)
    survey_question_family_count = len(code_to_cols) - operational_count

    group_candidates = detect_group_candidates(columns, data_rows)

    # respondent id duplicate check: first meta column with high uniqueness
    respondent_col = next((c for c in columns if c["is_meta"]), None)
    dup_count = None
    if respondent_col is not None:
        idx = respondent_col["index"] - 1
        vals = [row[idx] for row in data_rows if row[idx] not in (None, "")]
        dup_count = len(vals) - len(set(vals))

    profile = {
        "raw_path": str(raw_path),
        "response_n": len(data_rows),
        "column_n": max_col,
        "question_family_count": survey_question_family_count,
        "operational_field_count": operational_count,
        "family_question_type_counts": family_type_counts,
        "respondent_id_column": respondent_col["column_key"] if respondent_col else None,
        "respondent_id_duplicate_count": dup_count,
        "group_candidates": [
            {"column_key": g["column_key"], "n_levels": g["n_levels"], "levels": g["levels"]}
            for g in group_candidates
        ],
    }
    (out_dir / "dataset_profile.json").write_text(
        json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"response_n={profile['response_n']} column_n={profile['column_n']} "
          f"families={profile['question_family_count']} operational={operational_count} "
          f"types={family_type_counts}")


if __name__ == "__main__":
    main()
