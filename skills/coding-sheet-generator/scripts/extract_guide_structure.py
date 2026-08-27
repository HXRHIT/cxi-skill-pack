"""coding-sheet-generator canonical script (part 1 of 2 — guide extraction).

Reads an interview guide xlsx's `04. 인터뷰` and `형용사 테스트` sheets and
extracts the question structure a coding workbook needs: open-response
battery (Q2-7), adjective-card groups (Q8), recognition/usefulness pairs,
preference questions, and the numbered item lists for the two scale
questions (Q45 5-point, Q47 SUS 7-point).

Question numbers (2-7, 8, 15/19/20/23/40/41/42, 32/44/49, 45, 47) are specific
to this project's (`24.ST.GBI`) interview guide layout — a different guide
will need different target sets (`target_open_qs` etc. inside `main()`).
Each of those numbered questions does carry a matching `activity_type` label
in the guide itself (e.g. "인식과 유용함 간의 차이" for the recognition set,
"객관식-5점척도" for Q45) — deriving the target sets from that label instead
of a hardcoded number set would generalize better, but wasn't done in this
pass; the hardcoded sets were re-verified against the raw guide file and are
accurate for this project (see `04_validation_notes.md`).

Usage:
    python extract_guide_structure.py <interview_guide.xlsx> <out_dir>
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def normalize(text) -> str:
    if text is None:
        return ""
    return " ".join(str(text).replace("\r", " ").replace("\n", " ").split()).strip()


def parse_numbered_items(prompt: str) -> list[str]:
    items = []
    for line in str(prompt).splitlines():
        line = line.strip()
        if re.match(r"^\d+\.", line):
            items.append(re.sub(r"^\d+\.\s*", "", line))
    return items


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("guide_xlsx")
    parser.add_argument("out_dir")
    parser.add_argument("--project-id", default="unknown")
    args = parser.parse_args()

    guide_path = Path(args.guide_xlsx)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(guide_path, data_only=False)
    interview_ws = wb["04. 인터뷰"]
    adjective_ws = wb["형용사 테스트"]

    note_rows = []
    q2_q7 = []
    recognition_items = []
    preference_items = []
    q45_items = []
    q47_items = []

    target_open_qs = {2, 3, 4, 5, 6, 7}
    target_recognition_qs = {15, 19, 20, 23, 40, 41, 42}
    target_preference_qs = {32, 44, 49}

    for row_idx, row in enumerate(interview_ws.iter_rows(values_only=True), start=1):
        cols = list(row)
        question_no = cols[8] if len(cols) > 8 else None
        activity_type = normalize(cols[7] if len(cols) > 7 else "")
        note_record = {
            "row_index": row_idx,
            "stage": normalize(cols[4] if len(cols) > 4 else ""),
            "scenario": normalize(cols[5] if len(cols) > 5 else ""),
            "screen": normalize(cols[6] if len(cols) > 6 else ""),
            "activity_type": activity_type,
            "question_no": question_no if question_no not in ("", None) else "",
            "content": normalize(cols[9] if len(cols) > 9 else ""),
            "visual_asset": normalize(cols[10] if len(cols) > 10 else ""),
            "note": normalize(cols[11] if len(cols) > 11 else ""),
        }
        if any(note_record.values()):
            note_rows.append(note_record)

        if question_no in target_open_qs:
            q2_q7.append(
                {
                    "question_no": int(question_no),
                    "column_key": f"Q{int(question_no)}",
                    "screen": normalize(cols[6] if len(cols) > 6 else ""),
                    "stage": normalize(cols[4] if len(cols) > 4 else ""),
                    "prompt": normalize(cols[9] if len(cols) > 9 else ""),
                }
            )
        elif question_no == 8:
            adjective_prompt = {
                "question_no": 8,
                "prompt": normalize(cols[9] if len(cols) > 9 else ""),
                "activity_type": activity_type,
            }
        elif question_no in target_recognition_qs:
            recognition_items.append(
                {
                    "question_no": int(question_no),
                    "column_key": f"Q{int(question_no)}",
                    "screen": normalize(cols[6] if len(cols) > 6 else ""),
                    "stage": normalize(cols[4] if len(cols) > 4 else ""),
                    "prompt": normalize(cols[9] if len(cols) > 9 else ""),
                }
            )
        elif question_no in target_preference_qs:
            preference_items.append(
                {
                    "question_no": int(question_no),
                    "column_key": f"Q{int(question_no)}",
                    "screen": normalize(cols[6] if len(cols) > 6 else ""),
                    "stage": normalize(cols[4] if len(cols) > 4 else ""),
                    "prompt": normalize(cols[9] if len(cols) > 9 else ""),
                }
            )
        elif question_no == 45:
            # Bug fixed 2026-08-19: parse_numbered_items() splits on newlines
            # to find "N. item" lines, but normalize() collapses all
            # whitespace (including newlines) into single spaces. Calling it
            # on the *normalized* text always yielded 0 items — the raw cell
            # value must be parsed first, normalize() only for display.
            raw_q45 = cols[9] if len(cols) > 9 else ""
            q45_prompt = normalize(raw_q45)
            q45_items = parse_numbered_items(raw_q45)
        elif question_no == 47:
            raw_q47 = cols[9] if len(cols) > 9 else ""
            q47_prompt = normalize(raw_q47)
            q47_items = parse_numbered_items(raw_q47)

    adjective_groups = {"긍정": [], "중립": [], "부정": []}
    for row in adjective_ws.iter_rows(min_row=3, max_row=16, values_only=True):
        positive, neutral, negative = row
        if positive:
            adjective_groups["긍정"].append(normalize(positive))
        if neutral:
            adjective_groups["중립"].append(normalize(neutral))
        if negative:
            adjective_groups["부정"].append(normalize(negative))

    data = {
        "source_path": str(guide_path),
        "project_id": args.project_id,
        "guide_name": guide_path.name,
        "packaging_recommendation": "round_level_workbook",
        "note_rows": note_rows,
        "open_response_battery": sorted(q2_q7, key=lambda item: item["question_no"]),
        "adjective_card": {
            "question_no": 8,
            "prompt": adjective_prompt["prompt"],
            "groups": adjective_groups,
            "all_labels": adjective_groups["긍정"] + adjective_groups["중립"] + adjective_groups["부정"],
        },
        "recognition_usefulness": sorted(recognition_items, key=lambda item: item["question_no"]),
        "preference_questions": sorted(preference_items, key=lambda item: item["question_no"]),
        "q45_scale": {
            "question_no": 45,
            "scale_points": [1, 2, 3, 4, 5],
            "items": q45_items,
        },
        "q47_sus": {
            "question_no": 47,
            "scale_points": [1, 2, 3, 4, 5, 6, 7],
            "items": q47_items,
        },
        "reserved_participant_rows": 20,
        "participant_fields": [
            "PID",
            "이름/익명",
            "성별",
            "연령",
            "세션그룹",
            "인터뷰일시",
            "투자성향",
            "사전설문 메모",
            "코딩 메모",
        ],
    }

    (out_dir / "01_guide_structure.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
