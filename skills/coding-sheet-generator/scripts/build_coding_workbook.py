"""coding-sheet-generator canonical script (part 2 of 2 — workbook build).

Python/openpyxl port of the original `build_coding_workbook.mjs`, which
depended on a proprietary `@oai/artifact-tool` package not resolvable in this
environment (confirmed 2026-08-19 — `node -e "import('@oai/artifact-tool')"`
fails with MODULE_NOT_FOUND here). That import failure, not a bug in the
workbook logic itself, is almost certainly the real explanation for the
"bundled exporter ... non-zero exit" issue recorded in the 2026-08-18
validation notes: the script was very likely authored/run in a different
environment (e.g. Claude.ai's Artifacts tool) that does have that package,
and never ran cleanly in this agent-session environment at all.

This port reproduces the same 8-sheet structure, same formulas (translated
1:1, not re-derived), same dropdown validations, and same styling intent
(openpyxl's fill/font API differs in shape from the original but the same
colors and roles are used). It does not reproduce the PNG `renders/` step —
that relied on the proprietary tool's built-in sheet-to-image renderer, which
has no direct equivalent here; use the `xlsx` skill's LibreOffice-based
screenshot flow separately if visual renders are needed.

Usage:
    python build_coding_workbook.py <01_guide_structure.json> <out_dir>
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLORS = {
    "header": "0F766E",
    "headerText": "FFFFFF",
    "subHeader": "D9F99D",
    "paleBlue": "DBEAFE",
    "palePurple": "EDE9FE",
    "paleAmber": "FEF3C7",
    "paleGreen": "DCFCE7",
    "paleGray": "F3F4F6",
    "border": "D1D5DB",
    "text": "111827",
    "muted": "6B7280",
}

RESERVED_ROWS_DEFAULT = 20

DEFAULT_PARTICIPANT_FIELDS = [
    "PID", "이름/익명", "성별", "연령", "세션그룹", "인터뷰일시", "투자성향", "사전설문 메모", "코딩 메모",
]
DEFAULT_Q45_ITEMS = [
    "1. (간편설계의) 쉽고 간편한 설계", "2. (상세설계의) 상세한 설계", "3. 상세설계의 입력값 추천과 도움말",
    "4. 다른 금융사 연금 가져오기", "5. 다른 금융사와 합산하여 연금 계획 세우기", "6. 목표 (수정) 제안",
    "7. 포트폴리오 구성", "8. 맞춤 상품 추천", "9. 리밸런싱", "10. 설계 내역", "11. 투자수익률 알림",
]
DEFAULT_Q47_ITEMS = [
    "나는 이 서비스를 자주 사용할 것 같다", "나는 이 서비스가 불필요하게 복잡하다고 느꼈다",
    "나는 이 서비스는 사용하기 편리하다고 느꼈다", "내가 이 서비스를 사용하기 위해선 기술적 도움이 필요할 것 같다고 느꼈다",
    "나는 이 서비스에는 다양한 기능이 잘 통합되어있다고 생각했다", "나는 이 서비스에 일관적이지 않은 부분이 너무 많다고 느꼈다",
    "나는 대부분의 사람들이 이 서비스를 사용하는 방법을 쉽게 배울 것이라고 생각한다",
    "나는 이 서비스는 사용하기 불편하고 어색하다고 느꼈다", "나는 이 서비스를 사용하는 것에 자신감을 느꼈다",
    "이 서비스를 사용하기 전에 많은 것을 배워야 할 것 같다고 느꼈다",
]
DEFAULT_RECOGNITION_ITEMS = [
    {"question_no": 15, "screen": "목표 진단", "prompt": "목표를 수정할까요?"},
    {"question_no": 19, "screen": "상세 포트폴리오", "prompt": "상세 포트폴리오 보기"},
    {"question_no": 20, "screen": "ETF 추천 상품", "prompt": "ETF 상품도 있어요!"},
    {"question_no": 23, "screen": "설계 결과", "prompt": "투자는 교체 매매로 진행돼요!"},
    {"question_no": 40, "screen": "서비스 소개", "prompt": "설계 내역"},
    {"question_no": 41, "screen": "설계내역", "prompt": "투자수익률알림설정하기"},
    {"question_no": 42, "screen": "설계내역", "prompt": "스마트연금케어시작하기"},
]
DEFAULT_PREFERENCE_ITEMS = [
    {"question_no": 32, "prompt": "간편설계 vs 상세설계 중 더 선호하는 방식과 이유"},
    {"question_no": 44, "prompt": "설계내역 vs 타 앱 관리 방식 중 더 선호하는 방식과 이유"},
    {"question_no": 49, "prompt": "숫자 vs 이미지 표현 방식 중 더 선호하는 방식과 이유"},
]
DEFAULT_OPEN_BATTERY = [
    {"question_no": 2, "screen": "서비스 소개", "prompt": "가장 기억에 남는 키워드 3개는?"},
    {"question_no": 3, "screen": "설계 방식 선택", "prompt": "두 가지 방식의 차이는 무엇 같나요?"},
    {"question_no": 4, "screen": "상세설계", "prompt": "어떤 버튼을 눌렀을 것 같나요?"},
    {"question_no": 5, "screen": "목표 진단", "prompt": "가장 기억에 남는 정보는 무엇이었나요?"},
    {"question_no": 6, "screen": "AI 컨설팅", "prompt": "어떤 화면으로 이해했나요?"},
    {"question_no": 7, "screen": "설계 결과", "prompt": "다음 단계는 무엇이라고 생각하나요?"},
]
DEFAULT_ADJECTIVE_GROUPS = {
    "긍정": ["흥미진진한", "편리한", "친절한", "차분한", "사용하기 쉬운", "재미있는", "직관적인", "세련된", "인상적인", "전문적인", "고급스러운", "신뢰할 수 있는", "간결한", "유용한"],
    "중립": ["일관된", "명확한", "단순한", "익숙한", "최신의"],
    "부정": ["지루한", "일관성 없는", "산만한", "혼란스러운", "사용하기 어려운", "딱딱한", "복잡한", "세련되지 않은", "평범한", "품질이 낮은"],
}

THIN_BORDER = Border(*(Side(style="thin", color=COLORS["border"]) for _ in range(4)))


def cl(index: int) -> str:
    return get_column_letter(index)


def set_range_border(ws, cell_range: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.border = THIN_BORDER


def write_merged_title(ws, title: str, last_col: str) -> None:
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["header"])
    ws["A1"].font = Font(bold=True, color=COLORS["headerText"], size=15)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    set_range_border(ws, f"A1:{last_col}1")
    ws.row_dimensions[1].height = 28


def write_note_band(ws, text: str, last_col: str) -> None:
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = text
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["paleGray"])
    ws["A2"].font = Font(color=COLORS["text"], italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    set_range_border(ws, f"A2:{last_col}2")
    ws.row_dimensions[2].height = 34


def style_header(ws, cell_range: str, fill: str = COLORS["header"], font_color: str = COLORS["headerText"]) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(bold=True, color=font_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    set_range_border(ws, cell_range)


def style_body(ws, cell_range: str, horizontal: str = "left") -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(color=COLORS["text"])
            cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=True)
    set_range_border(ws, cell_range)


def set_column_widths(ws, pairs: list[tuple[str, int]]) -> None:
    for col, width in pairs:
        ws.column_dimensions[col].width = width


def write_participant_id_formulas(ws, start_row: int, reserved_rows: int, meta_start_row: int = 5) -> None:
    for i in range(reserved_rows):
        ws.cell(row=start_row + i, column=1, value=f"='01. 참여자 메타'!A{meta_start_row + i}")


def bool_dropdown(ws, cell_range: str) -> None:
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def list_dropdown(ws, cell_range: str, options: list[str]) -> None:
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def whole_range_dropdown(ws, cell_range: str, lo: int, hi: int, title: str, message: str) -> None:
    dv = DataValidation(
        type="whole", operator="between", formula1=str(lo), formula2=str(hi),
        allow_blank=True, showErrorMessage=True, errorTitle=title, error=message,
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_notes_sheet(wb, note_rows: list[dict]) -> None:
    ws = wb["00. 노트테이킹"]
    last_col = "H"
    write_merged_title(ws, "00. 노트테이킹", last_col)
    write_note_band(ws, "실제 세션 중 관찰/직접 인용/후속 탐침을 빠르게 적을 수 있도록 질문 단위 메모 행을 미리 배치했습니다.", last_col)

    headers = ["Stage", "Screen", "Activity Type", "Q#", "Prompt", "Observation / Behavior", "Key Quote", "Follow-up / Coding Memo"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, f"A4:{last_col}4")

    for r, row in enumerate(note_rows, start=5):
        ws.cell(row=r, column=1, value=row.get("stage", ""))
        ws.cell(row=r, column=2, value=row.get("screen", ""))
        ws.cell(row=r, column=3, value=row.get("activity_type", ""))
        ws.cell(row=r, column=4, value=row.get("question_no", ""))
        ws.cell(row=r, column=5, value=row.get("prompt") or row.get("content") or "")
    if note_rows:
        style_body(ws, f"A5:{last_col}{4 + len(note_rows)}")

    set_column_widths(ws, [("A", 18), ("B", 18), ("C", 18), ("D", 8), ("E", 55), ("F", 28), ("G", 28), ("H", 28)])
    ws.freeze_panes = "A5"


def build_meta_sheet(wb, participant_fields: list[str], reserved_rows: int) -> None:
    ws = wb["01. 참여자 메타"]
    headers = participant_fields + ["완료여부"]
    last_col = cl(len(headers))
    write_merged_title(ws, "01. 참여자 메타", last_col)
    write_note_band(ws, "참여자 메타는 다른 시트의 PID 기준값으로 연결됩니다. 이름 대신 익명 ID만 써도 됩니다.", last_col)

    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, f"A4:{last_col}4")

    for i in range(reserved_rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=f'="P"&TEXT(ROW()-4,"00")')
    style_body(ws, f"A5:{last_col}{4 + reserved_rows}")
    list_dropdown(ws, f"J5:J{4 + reserved_rows}", ["미시작", "진행중", "완료"])

    set_column_widths(ws, [("A", 10), ("B", 16), ("C", 10), ("D", 10), ("E", 14), ("F", 16), ("G", 12), ("H", 18), ("I", 18), ("J", 12)])
    ws.freeze_panes = "B5"


def build_open_battery_sheet(wb, open_battery: list[dict], reserved_rows: int) -> None:
    ws = wb["02. 주관식_10초테스트_Q2-Q7"]
    last_col = cl(len(open_battery) + 2)
    write_merged_title(ws, "02. 주관식_10초테스트_Q2-Q7", last_col)
    write_note_band(ws, "5초/10초 테스트형 짧은 자유응답을 한 행에 한 참여자씩 기록합니다. 마지막 열은 빠른 요약용입니다.", last_col)

    headers = ["ID"] + [f"Q{item['question_no']}\n[{item['screen']}]\n{item['prompt']}" for item in open_battery] + ["핵심 메모"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, f"A4:{last_col}4")
    ws.row_dimensions[4].height = 88

    write_participant_id_formulas(ws, 5, reserved_rows)
    style_body(ws, f"A5:{last_col}{4 + reserved_rows}")

    set_column_widths(ws, [("A", 10), ("B", 26), ("C", 26), ("D", 26), ("E", 26), ("F", 26), ("G", 26), ("H", 20)])
    ws.freeze_panes = "B5"


def build_adjective_sheet(wb, adjective_groups: dict, reserved_rows: int) -> None:
    ws = wb["03. 객관식-형용사카드_Q8"]
    all_labels = [label for labels in adjective_groups.values() for label in labels]
    last_col = cl(len(all_labels) + 2)
    write_merged_title(ws, "03. 객관식-형용사카드_Q8", last_col)
    write_note_band(ws, "참여자가 고른 형용사는 B열에 요약하고, 개별 카드 선택 여부는 TRUE/FALSE로 표시합니다. 6행은 빈도 확인용입니다.", last_col)

    ws["A5"], ws["B5"] = "ID", "선택 형용사(쉼표)"
    style_header(ws, "A4:B5", COLORS["paleGray"], COLORS["text"])

    group_fill = {"긍정": COLORS["paleGreen"], "중립": COLORS["paleBlue"], "부정": COLORS["paleAmber"]}
    current_col = 3
    for group_name, labels in adjective_groups.items():
        start, end = cl(current_col), cl(current_col + len(labels) - 1)
        ws.merge_cells(f"{start}4:{end}4")
        ws[f"{start}4"] = group_name
        style_header(ws, f"{start}4:{end}4", group_fill.get(group_name, COLORS["paleGray"]), COLORS["text"])
        current_col += len(labels)

    for i, label in enumerate(all_labels):
        ws.cell(row=5, column=3 + i, value=label)
    style_header(ws, f"C5:{last_col}5")
    ws.row_dimensions[5].height = 56

    ws["A6"], ws["B6"] = "count", "Q8 선택 빈도"
    style_header(ws, "A6:B6", COLORS["paleGray"], COLORS["text"])
    for i in range(len(all_labels)):
        col = cl(i + 3)
        ws.cell(row=6, column=3 + i, value=f"=COUNTIF({col}7:{col}{6 + reserved_rows},TRUE)")
    style_header(ws, f"C6:{last_col}6", COLORS["paleGray"], COLORS["text"])

    write_participant_id_formulas(ws, 7, reserved_rows)
    style_body(ws, f"A7:{last_col}{6 + reserved_rows}", "center")
    for r in range(7, 7 + reserved_rows):
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    bool_dropdown(ws, f"C7:{last_col}{6 + reserved_rows}")

    set_column_widths(ws, [("A", 10), ("B", 26)])
    for i in range(3, len(all_labels) + 3):
        ws.column_dimensions[cl(i)].width = 11
    ws.freeze_panes = "C7"


def build_recognition_sheet(wb, recognition_items: list[dict], reserved_rows: int) -> None:
    ws = wb["04. 인지와 유용_Q15-Q42"]
    total_cols = 1 + len(recognition_items) * 2 + 1
    last_col = cl(total_cols)
    write_merged_title(ws, "04. 인지와 유용_Q15-Q42", last_col)
    write_note_band(ws, "각 질문마다 참여자가 내용을 '인지'했는지와 '유용'하다고 느꼈는지를 TRUE/FALSE로 기록합니다. 마지막 열은 메모용입니다.", last_col)

    ws.cell(row=4, column=1, value="질문번호")
    ws.cell(row=5, column=1, value="질문 내용/스크린")
    ws.cell(row=6, column=1, value="ID")
    col = 2
    for item in recognition_items:
        ws.cell(row=4, column=col, value=item["question_no"])
        ws.cell(row=4, column=col + 1, value=item["question_no"])
        label = item.get("screen") or item.get("prompt") or ""
        ws.cell(row=5, column=col, value=label)
        ws.cell(row=5, column=col + 1, value=label)
        ws.cell(row=6, column=col, value="인지")
        ws.cell(row=6, column=col + 1, value="유용")
        col += 2
    ws.cell(row=6, column=col, value="종합 메모")

    style_header(ws, f"A4:{last_col}4", COLORS["palePurple"], COLORS["text"])
    style_header(ws, f"A5:{last_col}5", COLORS["paleBlue"], COLORS["text"])
    style_header(ws, f"A6:{last_col}6", COLORS["header"], COLORS["headerText"])
    ws.row_dimensions[5].height = 48

    write_participant_id_formulas(ws, 7, reserved_rows)
    style_body(ws, f"A7:{last_col}{6 + reserved_rows}", "center")
    for r in range(7, 7 + reserved_rows):
        ws.cell(row=r, column=total_cols).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    bool_last_col = cl(total_cols - 1)
    bool_dropdown(ws, f"B7:{bool_last_col}{6 + reserved_rows}")

    count_row = 7 + reserved_rows
    ws.cell(row=count_row, column=1, value="count_TRUE")
    for i in range(len(recognition_items) * 2):
        col_idx = i + 2
        col_letter = cl(col_idx)
        ws.cell(row=count_row, column=col_idx, value=f"=COUNTIF({col_letter}7:{col_letter}{6 + reserved_rows},TRUE)")
    style_header(ws, f"A{count_row}:{last_col}{count_row}", COLORS["paleGray"], COLORS["text"])

    set_column_widths(ws, [("A", 10)])
    for i in range(2, total_cols):
        ws.column_dimensions[cl(i)].width = 12
    ws.column_dimensions[last_col].width = 20
    ws.freeze_panes = "B7"


def build_preference_sheet(wb, preference_items: list[dict], reserved_rows: int) -> None:
    ws = wb["05. 선호_코딩_Q32_Q44_Q49"]
    last_col = "H"
    write_merged_title(ws, "05. 선호_코딩_Q32_Q44_Q49", last_col)
    write_note_band(ws, "선호 선택과 이유를 함께 남길 수 있게 설계했습니다. 드롭다운 값은 실제 코딩 시 필요하면 더 좁히거나 늘릴 수 있습니다.", last_col)

    items = (preference_items + [{}, {}, {}])[:3]
    header_prompts = [None] + [items[0].get("prompt", "")] * 2 + [items[1].get("prompt", "")] * 2 + [items[2].get("prompt", "")] * 2 + [None]
    headers = ["ID", "Q32 선택", "Q32 이유 요약", "Q44 선택", "Q44 이유 요약", "Q49 선택", "Q49 이유 요약", "종합 메모"]

    for i, v in enumerate(header_prompts, start=1):
        ws.cell(row=4, column=i, value=v)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    style_header(ws, "A4:H4", COLORS["paleBlue"], COLORS["text"])
    style_header(ws, "A5:H5")
    ws.row_dimensions[4].height = 60

    write_participant_id_formulas(ws, 6, reserved_rows, meta_start_row=5)
    style_body(ws, f"A6:H{5 + reserved_rows}")
    for r in range(6, 6 + reserved_rows):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    list_dropdown(ws, f"B6:B{5 + reserved_rows}", ["간편설계", "상세설계", "둘 다/기타"])
    list_dropdown(ws, f"D6:D{5 + reserved_rows}", ["설계내역", "타 앱 관리", "둘 다/기타"])
    list_dropdown(ws, f"F6:F{5 + reserved_rows}", ["숫자", "이미지", "둘 다/기타"])

    set_column_widths(ws, [("A", 10), ("B", 14), ("C", 22), ("D", 14), ("E", 22), ("F", 14), ("G", 22), ("H", 22)])
    ws.freeze_panes = "B6"


def build_q45_sheet(wb, q45_items: list[str], reserved_rows: int) -> None:
    ws = wb["06. 객관식-5점척도_Q45"]
    total_cols = 1 + len(q45_items) + 2
    last_col = cl(total_cols)
    write_merged_title(ws, "06. 객관식-5점척도_Q45", last_col)
    write_note_band(ws, "1~5점 응답을 참여자별 행 기준으로 입력합니다. 마지막 두 열은 평균과 응답 개수를 자동 계산합니다.", last_col)

    headers = ["ID"] + q45_items + ["평균", "응답 개수"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, f"A4:{last_col}4")
    ws.row_dimensions[4].height = 72

    write_participant_id_formulas(ws, 5, reserved_rows)
    last_item_col = cl(1 + len(q45_items))
    for i in range(reserved_rows):
        r = 5 + i
        ws.cell(row=r, column=1 + len(q45_items) + 1,
                value=f'=IF(COUNTA(B{r}:{last_item_col}{r})=0,"",ROUND(AVERAGE(B{r}:{last_item_col}{r}),2))')
        ws.cell(row=r, column=1 + len(q45_items) + 2, value=f"=COUNT(B{r}:{last_item_col}{r})")
    style_body(ws, f"A5:{last_col}{4 + reserved_rows}", "center")
    whole_range_dropdown(ws, f"B5:{last_item_col}{4 + reserved_rows}", 1, 5, "점수 범위 오류", "1점부터 5점 사이의 정수만 입력하세요.")

    summary_count_row = 5 + reserved_rows
    summary_mean_row = 6 + reserved_rows
    ws.cell(row=summary_count_row, column=1, value="응답수")
    ws.cell(row=summary_mean_row, column=1, value="항목 평균")
    for i in range(len(q45_items)):
        col_letter = cl(i + 2)
        ws.cell(row=summary_count_row, column=i + 2, value=f"=COUNT({col_letter}5:{col_letter}{4 + reserved_rows})")
        ws.cell(row=summary_mean_row, column=i + 2,
                value=f'=IF(COUNT({col_letter}5:{col_letter}{4 + reserved_rows})=0,"",ROUND(AVERAGE({col_letter}5:{col_letter}{4 + reserved_rows}),2))')
    style_header(ws, f"A{summary_count_row}:{last_col}{summary_mean_row}", COLORS["paleGray"], COLORS["text"])

    set_column_widths(ws, [("A", 10)])
    for i in range(2, 13):
        ws.column_dimensions[cl(i)].width = 16
    ws.freeze_panes = "B5"


def build_q47_sheet(wb, q47_items: list[str], reserved_rows: int) -> None:
    ws = wb["07. 객관식-7점척도_Q47_SUS"]
    total_cols = 1 + len(q47_items) + 3
    last_col = cl(total_cols)
    write_merged_title(ws, "07. 객관식-7점척도_Q47_SUS", last_col)
    write_note_band(ws, "1~7점 응답을 입력하면 SUS 점수(100점 만점)와 Grade를 자동 계산합니다. 홀수 문항은 정방향, 짝수 문항은 역방향으로 계산합니다.", last_col)

    headers = ["ID"] + [f"Q{i + 1}\n{item}" for i, item in enumerate(q47_items)] + ["SUS 점수", "SUS Grade", "비고"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, f"A4:{last_col}4")
    ws.row_dimensions[4].height = 96

    write_participant_id_formulas(ws, 5, reserved_rows)
    # SUS score: 10-item scale, odd items (1,3,5,7,9 -> cols B,D,F,H,J) score (x-1),
    # even items (2,4,6,8,10 -> cols C,E,G,I,K) score (7-x).
    #
    # Bug found and fixed 2026-08-19: the original .mjs script multiplied this
    # sum by 2.5 — correct for the *classic 5-point* SUS scale (each item
    # contributes 0-4, so 10 items max out at 40, and 40*2.5=100), but this
    # workbook uses a 1-7 point scale (each item contributes 0-6, 10 items max
    # out at 60). 2.5 on a 1-7 scale scores "all neutral" as 75 instead of 50
    # and lets a maximally-positive respondent score 150 — over the 0-100
    # range the SUS grade thresholds below assume. The correct multiplier for
    # a 10-item 1-7 scale is 100/60 = 5/3, verified by hand: neutral (all 4s)
    # -> 50, best case -> 100, worst case -> 0. This was almost certainly
    # never caught because the original pipeline's `@oai/artifact-tool`
    # dependency doesn't resolve in this environment (see module docstring),
    # so nobody had opened a working copy of this sheet to notice the range
    # was off. Filed here as a fixed bug, not a preserved-for-fidelity quirk.
    for i in range(reserved_rows):
        r = 5 + i
        terms = []
        for item_idx in range(len(q47_items)):
            col_letter = cl(item_idx + 2)
            if item_idx % 2 == 0:
                terms.append(f"({col_letter}{r}-1)")
            else:
                terms.append(f"(7-{col_letter}{r})")
        last_item_col = cl(1 + len(q47_items))
        # *(5/3) written as a literal formula fraction, not a rounded decimal,
        # so Excel/LibreOffice computes the exact value.
        formula = f'=IF(COUNTA(B{r}:{last_item_col}{r})=0,"",(' + "+".join(terms) + ")*(5/3))"
        sus_col = 1 + len(q47_items) + 1
        grade_col = sus_col + 1
        remark_col = grade_col + 1
        ws.cell(row=r, column=sus_col, value=formula)
        sus_cell = cl(sus_col)
        ws.cell(row=r, column=grade_col,
                value=(f'=IF({sus_cell}{r}="","",IF({sus_cell}{r}>=80.3,"A",'
                       f'IF({sus_cell}{r}>=68,"B",IF({sus_cell}{r}>=51,"C",'
                       f'IF({sus_cell}{r}>=38,"D","F")))))'))
        ws.cell(row=r, column=remark_col, value=None)

    last_item_col = cl(1 + len(q47_items))
    style_body(ws, f"A5:{last_col}{4 + reserved_rows}", "center")
    whole_range_dropdown(ws, f"B5:{last_item_col}{4 + reserved_rows}", 1, 7, "점수 범위 오류", "1점부터 7점 사이의 정수만 입력하세요.")
    remark_col_letter = cl(1 + len(q47_items) + 3)
    for r in range(5, 5 + reserved_rows):
        ws.cell(row=r, column=1 + len(q47_items) + 3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    summary_count_row = 5 + reserved_rows
    summary_mean_row = 6 + reserved_rows
    ws.cell(row=summary_count_row, column=1, value="응답수")
    ws.cell(row=summary_mean_row, column=1, value="항목 평균")
    for i in range(len(q47_items)):
        col_letter = cl(i + 2)
        ws.cell(row=summary_count_row, column=i + 2, value=f"=COUNT({col_letter}5:{col_letter}{4 + reserved_rows})")
        ws.cell(row=summary_mean_row, column=i + 2,
                value=f'=IF(COUNT({col_letter}5:{col_letter}{4 + reserved_rows})=0,"",ROUND(AVERAGE({col_letter}5:{col_letter}{4 + reserved_rows}),2))')
    sus_col_letter = cl(1 + len(q47_items) + 1)
    ws.cell(row=summary_count_row, column=1 + len(q47_items) + 1, value=f"=COUNT({sus_col_letter}5:{sus_col_letter}{4 + reserved_rows})")
    ws.cell(row=summary_mean_row, column=1 + len(q47_items) + 1,
            value=f'=IF(COUNT({sus_col_letter}5:{sus_col_letter}{4 + reserved_rows})=0,"",ROUND(AVERAGE({sus_col_letter}5:{sus_col_letter}{4 + reserved_rows}),2))')
    style_header(ws, f"A{summary_count_row}:{last_col}{summary_mean_row}", COLORS["paleGray"], COLORS["text"])

    set_column_widths(ws, [("A", 10)])
    for i in range(2, 12):
        ws.column_dimensions[cl(i)].width = 16
    ws.freeze_panes = "B5"


ORDERED_SHEET_NAMES = [
    "00. 노트테이킹", "01. 참여자 메타", "02. 주관식_10초테스트_Q2-Q7", "03. 객관식-형용사카드_Q8",
    "04. 인지와 유용_Q15-Q42", "05. 선호_코딩_Q32_Q44_Q49", "06. 객관식-5점척도_Q45", "07. 객관식-7점척도_Q47_SUS",
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("guide_structure_json")
    parser.add_argument("out_dir")
    args = parser.parse_args()

    structure = json.loads(Path(args.guide_structure_json).read_text(encoding="utf-8"))
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    reserved_rows = structure.get("reserved_participant_rows", RESERVED_ROWS_DEFAULT)
    participant_fields = structure.get("participant_fields") or DEFAULT_PARTICIPANT_FIELDS
    q45_items = (structure.get("q45_scale") or {}).get("items") or DEFAULT_Q45_ITEMS
    q47_items = (structure.get("q47_sus") or {}).get("items") or DEFAULT_Q47_ITEMS
    recognition_items = structure.get("recognition_usefulness") or DEFAULT_RECOGNITION_ITEMS
    preference_items = structure.get("preference_questions") or DEFAULT_PREFERENCE_ITEMS
    open_battery = structure.get("open_response_battery") or DEFAULT_OPEN_BATTERY
    adjective_groups = (structure.get("adjective_card") or {}).get("groups") or DEFAULT_ADJECTIVE_GROUPS

    targeted_questions = {item["question_no"] for item in open_battery}
    targeted_questions.add(8)
    targeted_questions.update(item["question_no"] for item in recognition_items)
    targeted_questions.update(item["question_no"] for item in preference_items)
    targeted_questions.update([45, 47])
    note_rows = [
        row for row in structure.get("note_rows", [])
        if isinstance(row.get("question_no"), (int, float)) and row["question_no"] in targeted_questions
    ]

    wb = Workbook()
    wb.remove(wb.active)
    for name in ORDERED_SHEET_NAMES:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False

    build_notes_sheet(wb, note_rows)
    build_meta_sheet(wb, participant_fields, reserved_rows)
    build_open_battery_sheet(wb, open_battery, reserved_rows)
    build_adjective_sheet(wb, adjective_groups, reserved_rows)
    build_recognition_sheet(wb, recognition_items, reserved_rows)
    build_preference_sheet(wb, preference_items, reserved_rows)
    build_q45_sheet(wb, q45_items, reserved_rows)
    build_q47_sheet(wb, q47_items, reserved_rows)

    output_path = out_dir / "coding-workbook.xlsx"
    wb.save(output_path)

    verification_summary = {
        "source_path": structure.get("source_path"),
        "guide_name": structure.get("guide_name"),
        "generated_by": "build_coding_workbook.py (openpyxl port, 2026-08-19)",
        "fallback_usage": {
            "q45_items_from_fallback": not bool((structure.get("q45_scale") or {}).get("items")),
            "q47_items_from_fallback": not bool((structure.get("q47_sus") or {}).get("items")),
        },
        "sheets": ORDERED_SHEET_NAMES,
        "note": "PNG renders and workbook.inspect()-style formula snapshots from the "
                "original @oai/artifact-tool pipeline are not reproduced by this port "
                "— run recalc.py separately to verify formulas evaluate cleanly.",
    }
    (out_dir / "03_verification_summary.json").write_text(
        json.dumps(verification_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"wrote {output_path} with sheets: {', '.join(ORDERED_SHEET_NAMES)}")


if __name__ == "__main__":
    main()
