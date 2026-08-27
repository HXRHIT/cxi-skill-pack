#!/usr/bin/env python3
"""전사본 정확도 검증·정제 파이프라인 (canonical script for `transcript-verification-enhancer`).

2026-08-27 1차 검증에서 Critical 6건이 확인되어 재작성했다.
검증 근거: `validation_runs/transcript-verification-enhancer/2026-08-27_native-transcripts/`
구버전 원본은 같은 폴더의 `batch_processor_before_fix.py`.

## 설계 원칙 (검증에서 나온 것)

1. **행 구조를 바꾸지 않는다.** 구버전은 같은 화자의 연속 행을 병합하면서 `Context` 외
   컬럼을 `'first'`로 집계해 값을 버렸다 — IBUJA 코딩 워크북에서 `질문` 294셀, `시간` 291셀,
   `Code` 41셀이 삭제됐다(`Code` 순감 35). 이제 병합은 **프롬프트 문맥용으로만** 하고
   결과는 원래 행에 되쓴다. 행 수는 입력과 항상 같다.
2. **원본 워크북을 통째로 보존한다.** 구버전은 `df.to_excel(sheet_name="Cleaned")`로
   7시트+차트1 → 1시트+차트0으로 만들었다. 이제 openpyxl로 원본을 열어 대상 셀만
   갱신하고 저장하므로 다른 시트·차트·서식이 남는다.
3. **시트·컬럼을 실측해서 찾는다.** 구버전 기본값 `--sheet "1. 녹취 종합"`은 native 전사본
   36건 중 1건에만 있었고(그 1건도 전사본이 아니라 코딩 워크북), 나머지 35건은 `Contents`다.
   `Contents`는 헤더행이 없고 배치가 `[빈칸, 화자, 타임스탬프, 발화]`라서 위치 가정
   (`columns[2]`=PID, `columns[4]`=Context)이 깨졌다.
4. **native 안에 쓰지 않는다.** 구버전은 출력을 입력과 같은 디렉터리에 만들어, native
   경로를 주면 native에 파일이 생겼다(AGENTS.md §0 위반).
5. **외부 전송은 명시적 동의 없이는 거부한다.** 이 스크립트는 내부 전사본 원문을 외부 LLM
   API로 보낸다. 조직 지침상 사전 협의가 필요하므로 `--i-have-upload-approval` 없이는
   LLM 호출을 하지 않는다.
6. **익명화는 이 스킬이 하지 않는다.** `$transcript-anonymizer-skill`(#14)이 담당하며
   그쪽은 치환 전 사람 승인을 요구한다. 여기서 인라인 치환하면 그 게이트를 우회한다.
   구버전 프롬프트는 `P7 -> [P001]` 재치환까지 지시했는데, 이는 #14 검증에서 Critical로
   잡힌 "이미 익명화된 라벨 재치환"이다.

## 사용

    # 1) 구조만 확인 (읽기만, 쓰기 없음)
    python batch_processor.py IN.xlsx --inspect

    # 2) LLM 없이 파이프라인만 실행 (원문 유지, change log 생성)
    python batch_processor.py IN.xlsx --out OUT.xlsx --dry-run

    # 3) 실제 정제 (외부 API 전송 사전 협의 필요)
    python batch_processor.py IN.xlsx --out OUT.xlsx --i-have-upload-approval

## 가정 (새 데이터에 쓰기 전에 확인)

- .xlsx / .xlsm / .csv 지원. `.xls`는 `xlrd`가 필요하고 이 환경에는 없다.
- 발화 컬럼 판정은 "평균 길이가 가장 긴 텍스트 컬럼"이다. 발화보다 긴 메모 컬럼이 있으면
  `--text-col`로 직접 지정해야 한다.
"""
import argparse
import json
import re
import sys
import time
from pathlib import Path

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

SPEAKER_HINTS = ("진행자", "참석자", "참여자", "모더레이터", "사회자", "연구원")
TIME_RX = re.compile(r"^\s*\d{1,2}:\d{2}(:\d{2})?\s*$")
NAMED_HEADERS = ("PID", "Context", "질문", "시간", "Code", "Insight", "#(자동)")

SYSTEM_PROMPT = """You are cleaning UX research interview transcripts.

Apply exactly these four rules and nothing else:
1. Anaphora resolution - clarify ambiguous pronouns by appending the actual referent in parentheses, inferred from surrounding turns.
2. Turn separation - if a question and an answer are clumped into one row, separate them with a newline and a speaker marker.
3. Domain typo correction - fix UX/finance vocabulary misrecognitions. Never guess or change a person's name.
4. Filler removal - drop non-semantic fillers and immediate duplications.

Hard constraints:
- Do NOT anonymize. Leave names, labels, company names, and service names exactly as they are. Anonymization is a separate, human-approved step handled by another skill.
- Do NOT rewrite, summarize, translate, reorder, or invent content. Preserve conversational nuance.
- Return exactly one output row per input row, with the same id. Never merge, split, or drop rows.

Input format: JSON array of {id, speaker, context, neighbors}
Output format: JSON array of {id, context}. RAW JSON only, no markdown fence.
"""


# ------------------------------------------------------------------ 구조 탐지

def pick_sheet(sheetnames, explicit):
    """시트명을 실측으로 고른다. 구버전의 하드코딩 기본값을 대체한다."""
    if explicit:
        if explicit not in sheetnames:
            raise SystemExit("[중단] 시트 '%s' 가 없다. 이 파일의 시트: %s" % (explicit, sheetnames))
        return explicit
    for cand in ("Contents", "1. 녹취 종합", "녹취 종합"):
        if cand in sheetnames:
            return cand
    if len(sheetnames) == 1:
        return sheetnames[0]
    raise SystemExit("[중단] 어느 시트를 처리할지 판단할 수 없다. --sheet 로 지정하라. 시트: %s"
                     % sheetnames)


def profile_columns(rows, max_scan=200):
    """행 리스트에서 화자·타임스탬프·발화 컬럼 인덱스를 추정한다."""
    ncol = max((len(r) for r in rows[:max_scan]), default=0)
    if ncol == 0:
        raise SystemExit("[중단] 빈 시트다.")

    stats = []
    for c in range(ncol):
        texts = [str(r[c]) for r in rows[:max_scan]
                 if c < len(r) and r[c] is not None and str(r[c]).strip() != ""]
        stats.append({
            "idx": c,
            "n": len(texts),
            "avg_len": (sum(len(t) for t in texts) / len(texts)) if texts else 0,
            "uniq": len(set(texts)),
            "time_ratio": (sum(1 for t in texts if TIME_RX.match(t)) / len(texts)) if texts else 0,
            "hint_ratio": (sum(1 for t in texts if any(h in t for h in SPEAKER_HINTS))
                           / len(texts)) if texts else 0,
        })

    time_col = next((s["idx"] for s in stats if s["time_ratio"] > 0.6 and s["n"] > 3), None)

    cands = [s for s in stats if s["n"] > 3 and s["idx"] != time_col]
    speaker_col = next((s["idx"] for s in sorted(cands, key=lambda s: -s["hint_ratio"])
                        if s["hint_ratio"] > 0.3), None)
    if speaker_col is None:
        short = [s for s in cands if s["avg_len"] <= 12 and s["uniq"] <= 12]
        speaker_col = max(short, key=lambda s: s["n"])["idx"] if short else None

    pool = [s for s in stats if s["idx"] not in (time_col, speaker_col) and s["n"] > 0]
    if not pool:
        raise SystemExit("[중단] 발화 컬럼을 찾지 못했다. --text-col 로 지정하라.")
    text_col = max(pool, key=lambda s: s["avg_len"])["idx"]

    return {"speaker": speaker_col, "time": time_col, "text": text_col, "ncol": ncol}


def header_names(rows, ncol):
    """헤더행이 있으면 컬럼명을, 없으면 위치 라벨을 돌려준다(보고용)."""
    if not rows:
        return []
    first = rows[0]
    if any(c is not None and str(c).strip() in NAMED_HEADERS for c in first):
        return [None if c is None else str(c) for c in first]
    return ["(col%d)" % i for i in range(ncol)]


# ------------------------------------------------------------------ LLM

def call_llm(payload, model_name):
    import litellm
    for attempt in range(3):
        try:
            resp = litellm.completion(
                model=model_name,
                messages=[{"role": "system", "content": SYSTEM_PROMPT},
                          {"role": "user", "content": json.dumps(payload, ensure_ascii=False)}],
                temperature=0.1,
            )
            content = resp.choices[0].message.content.strip()
            content = re.sub(r"^```(?:json)?|```$", "", content).strip()
            return json.loads(content)
        except Exception as e:
            if type(e).__name__ == "RateLimitError" and attempt < 2:
                wait = 15 * (attempt + 1)
                print("  rate limit - %d초 대기 후 재시도" % wait)
                time.sleep(wait)
                continue
            print("  [경고] 청크 처리 실패(%s: %s) - 이 청크는 원문을 유지한다"
                  % (type(e).__name__, str(e)[:160]))
            return None
    return None


# ------------------------------------------------------------------ 본체

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input_file")
    ap.add_argument("--out", help="출력 파일 경로. native 안으로는 쓸 수 없다")
    ap.add_argument("--sheet", help="처리할 시트명. 미지정 시 실측으로 고른다")
    ap.add_argument("--text-col", type=int, help="발화 컬럼 인덱스(0-based) 직접 지정")
    ap.add_argument("--chunk-size", type=int, default=30)
    ap.add_argument("--model", default="anthropic/claude-sonnet-5")
    ap.add_argument("--inspect", action="store_true", help="구조만 보고 종료(쓰기 없음)")
    ap.add_argument("--dry-run", action="store_true",
                    help="LLM을 호출하지 않고 파이프라인만 실행(원문 유지, change log 생성)")
    ap.add_argument("--i-have-upload-approval", action="store_true",
                    help="내부 전사본 원문의 외부 LLM API 전송에 대한 사전 협의를 마쳤음")
    a = ap.parse_args()

    src = Path(a.input_file)
    if not src.exists():
        raise SystemExit("[중단] 입력 파일이 없다: %s" % src)

    ext = src.suffix.lower()
    if ext == ".xls":
        raise SystemExit("[중단] .xls 는 xlrd 가 필요하고 이 환경에 없다. .xlsx 로 변환해 오라.")

    wb = None
    sheet = None
    if ext == ".csv":
        import csv as _csv
        with open(src, encoding="utf-8-sig", newline="") as f:
            rows = [tuple(r) for r in _csv.reader(f)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(src, data_only=False)
        sheet = pick_sheet(wb.sheetnames, a.sheet)
        rows = [tuple(r) for r in wb[sheet].iter_rows(values_only=True)]

    prof = profile_columns(rows)
    if a.text_col is not None:
        prof["text"] = a.text_col
    names = header_names(rows, prof["ncol"])

    def label(i):
        if i is None:
            return "(없음)"
        return names[i] if i < len(names) and names[i] else "(col%d)" % i

    print("입력      : %s" % src.name)
    print("시트      : %s%s" % (sheet or "(csv)", "" if a.sheet else "   <- 실측으로 선택"))
    print("전체 시트 : %s" % (wb.sheetnames if wb else "-"))
    print("크기      : %d행 x %d열" % (len(rows), prof["ncol"]))
    print("화자 컬럼 : %s -> %s" % (prof["speaker"], label(prof["speaker"])))
    print("시간 컬럼 : %s -> %s" % (prof["time"], label(prof["time"])))
    print("발화 컬럼 : %s -> %s" % (prof["text"], label(prof["text"])))

    if a.inspect:
        return 0

    if not a.out:
        raise SystemExit("[중단] --out 이 필요하다. 원본 옆에 자동 생성하지 않는다.")
    out = Path(a.out).resolve()
    if "native" in out.parts:
        raise SystemExit("[중단] native repo 안으로는 쓸 수 없다(AGENTS.md §0): %s" % out)
    if out == src.resolve():
        raise SystemExit("[중단] 원본을 덮어쓸 수 없다.")

    if not a.dry_run and not a.i_have_upload_approval:
        raise SystemExit(
            "[중단] 이 스크립트는 전사본 원문을 외부 LLM API(%s)로 전송한다.\n"
            "       내부 자료 외부 전송은 사전 협의가 필요하다(조직 지침).\n"
            "       구조·파이프라인만 확인하려면 --dry-run,\n"
            "       협의를 마쳤으면 --i-have-upload-approval 을 붙여라." % a.model)

    text_i, spk_i = prof["text"], prof["speaker"]
    units = []
    for ri, r in enumerate(rows):
        val = r[text_i] if text_i < len(r) else None
        if val is None or str(val).strip() == "":
            continue
        spk = ""
        if spk_i is not None and spk_i < len(r) and r[spk_i] is not None:
            spk = str(r[spk_i]).strip()
        units.append({"row": ri, "speaker": spk, "text": str(val)})

    print("처리 대상 : %d행 (빈 발화 제외)" % len(units))
    if not units:
        raise SystemExit("[중단] 처리할 발화가 없다.")

    print("MODE      : %s" % ("dry-run - LLM 호출 없음, 원문 유지" if a.dry_run
                              else "실제 정제 (%s)" % a.model))

    changes = []
    for i in range(0, len(units), a.chunk_size):
        chunk = units[i:i + a.chunk_size]
        payload = []
        for j, u in enumerate(chunk):
            lo = max(0, i + j - 1)
            neighbors = [v["text"][:200] for v in units[lo:i + j + 2] if v is not u]
            payload.append({"id": u["row"], "speaker": u["speaker"],
                            "context": u["text"], "neighbors": neighbors})
        print("  %d-%d행 처리" % (chunk[0]["row"] + 1, chunk[-1]["row"] + 1))
        result = None if a.dry_run else call_llm(payload, a.model)
        if not result:
            continue
        by_id = {u["row"]: u for u in chunk}
        for item in result:
            rid, new = item.get("id"), item.get("context")
            if rid in by_id and isinstance(new, str) and new != by_id[rid]["text"]:
                changes.append({"row": rid + 1, "speaker": by_id[rid]["speaker"],
                                "before": by_id[rid]["text"], "after": new})
                by_id[rid]["new"] = new

    # 수식 캐시 경고: openpyxl은 수식은 보존하지만 **계산된 캐시값은 버린다.**
    # 실측(IBUJA 코딩 워크북) — 수식 5747셀은 그대로 남고, data_only로 읽히는 값은
    # 16542 -> 10811로 줄었다(`#(자동)` 열이 2533 -> 0). 하위 도구가 data_only=True로
    # 읽으면 빈 값을 본다. Excel에서 한 번 열어 재계산하거나 값고정이 필요하다.
    formula_cells = 0
    if wb is not None:
        for _ws in wb.worksheets:
            for _row in _ws.iter_rows():
                for _c in _row:
                    if isinstance(_c.value, str) and _c.value.startswith("="):
                        formula_cells += 1

    applied = 0
    if ext == ".csv":
        import csv as _csv
        mutable = [list(r) for r in rows]
        for u in units:
            if "new" in u:
                mutable[u["row"]][text_i] = u["new"]
                applied += 1
        with open(out, "w", encoding="utf-8-sig", newline="") as f:
            _csv.writer(f).writerows(mutable)
    else:
        ws = wb[sheet]
        for u in units:
            if "new" in u:
                ws.cell(row=u["row"] + 1, column=text_i + 1, value=u["new"])
                applied += 1
        wb.save(out)

    log_path = out.with_name(out.stem + "_changelog.md")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("# Change Log - %s\n\n" % out.name)
        f.write("| 항목 | 값 |\n|---|---|\n")
        f.write("| 입력 | `%s` |\n" % src.name)
        f.write("| 시트 | %s |\n" % (sheet or "(csv)"))
        f.write("| 입력 행 수 | %d |\n" % len(rows))
        f.write("| 출력 행 수 | %d |\n" % len(rows))
        f.write("| 처리 대상 발화 | %d |\n" % len(units))
        f.write("| 변경된 발화 | %d |\n" % applied)
        f.write("| 모드 | %s |\n" % ("dry-run (LLM 미호출)" if a.dry_run else a.model))
        f.write("| 보존된 시트 | %s |\n" % (", ".join(wb.sheetnames) if wb else "-"))
        f.write("| 수식 셀 | %d |\n" % formula_cells)
        f.write("\n행 수는 입력과 출력이 항상 같다 - 이 스크립트는 행을 병합하거나 삭제하지 않는다.\n")
        if formula_cells:
            f.write("\n> **수식 캐시 주의**: 수식 %d개는 그대로 보존됐지만, openpyxl 저장 과정에서 "
                    "**계산된 캐시값이 사라진다.** 하위 도구가 `data_only=True`로 읽으면 해당 셀이 "
                    "비어 보인다. Excel에서 한 번 열어 재계산하거나 값으로 고정한 뒤 넘겨라"
                    "(native 지침의 '구글시트 함수는 값고정'과 같은 사안).\n" % formula_cells)
        if changes:
            f.write("\n## 변경 내역\n\n| 행 | 화자 | 이전 | 이후 |\n|---|---|---|---|\n")
            for c in changes[:500]:
                f.write("| %d | %s | %s | %s |\n" % (
                    c["row"], c["speaker"],
                    c["before"][:160].replace("|", "\\|").replace("\n", " "),
                    c["after"][:160].replace("|", "\\|").replace("\n", " ")))
            if len(changes) > 500:
                f.write("\n(총 %d건 중 500건만 표기)\n" % len(changes))
        else:
            f.write("\n변경된 발화가 없다.\n")

    print()
    print("출력      : %s" % out)
    print("change log: %s" % log_path)
    print("행 수     : 입력 %d -> 출력 %d (동일)" % (len(rows), len(rows)))
    print("변경 발화 : %d" % applied)
    if wb:
        print("보존 시트 : %d개 %s" % (len(wb.sheetnames), wb.sheetnames))
    if formula_cells:
        print("[주의]    수식 %d개는 보존되지만 계산 캐시값은 사라진다 - "
              "Excel에서 한 번 열어 재계산하거나 값으로 고정한 뒤 넘길 것" % formula_cells)
    return 0


if __name__ == "__main__":
    sys.exit(main())
