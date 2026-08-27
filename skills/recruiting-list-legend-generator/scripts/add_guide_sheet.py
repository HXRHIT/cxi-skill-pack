#!/usr/bin/env python3
"""Add a leading `00. 사용가이드` sheet to a recruiting / participant-profile workbook.

canonical script for `recruiting-list-legend-generator` (AGENTS.md §4-1).

가정 (새 프로젝트에 쓰기 전에 확인할 것):
- 입력은 .xlsx (xlsm/xls 미지원). openpyxl로 읽고 다시 저장하므로 **차트·이미지·피벗·조건부서식 일부는 유실된다.**
  따라서 원본을 덮어쓰지 않고 항상 새 파일로 쓴다. 원본이 시각 자산을 담고 있으면 이 스크립트를 쓰지 말고
  가이드 시트만 별도 파일로 만들어 붙이는 경로를 택한다(--guide-only).
- guide 내용은 이 스크립트가 판단하지 않는다. JSON spec으로 받는다 — 워크북 해석은 사람/모델의 몫이고,
  이 스크립트는 렌더링과 시트 순서, 검증만 담당한다.

사용:
  python add_guide_sheet.py --src IN.xlsx --spec spec.json --out OUT.xlsx [--guide-only]
  python add_guide_sheet.py --src IN.xlsx --inspect-only          # 구조만 덤프

spec.json 스키마:
  {
    "title": "...",                        # 가이드 시트 H1
    "sections": [                           # 순서대로 렌더
      {"heading": "이 파일은 무엇인가", "rows": [["라벨","설명"], ...]},
      ...
    ]
  }
"""
import argparse, json, re, sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

GUIDE_SHEET = "00. 사용가이드"
ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def inspect(path):
    """구조 + 수식 오류 스캔. data_only 양쪽으로 두 번 읽는다."""
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    out = {"path": str(path), "sheets": [], "formula_errors": [], "formula_cells": 0}
    for ws in wb_f.worksheets:
        vs = wb_v[ws.title]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        out["sheets"].append({
            "index": wb_f.worksheets.index(ws),
            "title": ws.title,
            "dims": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "headers": [None if h is None else str(h) for h in headers],
        })
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    out["formula_cells"] += 1
        for row in vs.iter_rows():
            for c in row:
                if isinstance(c.value, str) and any(t in c.value for t in ERROR_TOKENS):
                    out["formula_errors"].append(
                        {"sheet": ws.title, "cell": c.coordinate, "value": c.value})
    return out


def write_guide(ws, spec):
    H1 = Font(bold=True, size=14)
    H2 = Font(bold=True, size=11)
    LBL = Font(bold=True)
    WRAP = Alignment(vertical="top", wrap_text=True)
    BAND = PatternFill("solid", fgColor="F2F2F2")

    r = 1
    ws.cell(r, 1, spec["title"]).font = H1
    r += 2
    for sec in spec["sections"]:
        c = ws.cell(r, 1, sec["heading"])
        c.font = H2
        c.fill = BAND
        ws.cell(r, 2, "").fill = BAND
        r += 1
        for row in sec["rows"]:
            ws.cell(r, 1, row[0]).font = LBL
            ws.cell(r, 1).alignment = WRAP
            if len(row) > 1:
                ws.cell(r, 2, row[1]).alignment = WRAP
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A3"
    return r


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True)
    ap.add_argument("--spec")
    ap.add_argument("--out")
    ap.add_argument("--inspect-only", action="store_true")
    ap.add_argument("--guide-only", action="store_true",
                    help="원본을 재저장하지 않고 가이드 시트만 담은 새 워크북을 만든다(시각 자산 유실 회피)")
    ap.add_argument("--verify-out")
    a = ap.parse_args()

    before = inspect(a.src)
    if a.inspect_only:
        json.dump(before, sys.stdout, ensure_ascii=False, indent=2)
        return

    spec = json.load(open(a.spec, encoding="utf-8"))

    if a.guide_only:
        wb = Workbook()
        ws = wb.active
        ws.title = GUIDE_SHEET
    else:
        wb = load_workbook(a.src, data_only=False)
        if GUIDE_SHEET in wb.sheetnames:
            del wb[GUIDE_SHEET]
        ws = wb.create_sheet(GUIDE_SHEET)
        # 선두 탭으로 재배치 — openpyxl은 _sheets 리스트를 직접 다룬다
        wb._sheets.remove(ws)
        wb._sheets.insert(0, ws)
        wb.active = 0

    write_guide(ws, spec)
    wb.save(a.out)

    after = inspect(a.out)
    verify = {
        "src": before,
        "out": after,
        "checks": {
            "guide_is_first_sheet": after["sheets"][0]["title"] == GUIDE_SHEET,
            "guide_sheet_present": any(s["title"] == GUIDE_SHEET for s in after["sheets"]),
            "source_sheets_preserved": (
                [s["title"] for s in before["sheets"]]
                == [s["title"] for s in after["sheets"] if s["title"] != GUIDE_SHEET]
            ),
            "row_counts_preserved": {
                s["title"]: {"before": s["max_row"],
                             "after": next((t["max_row"] for t in after["sheets"]
                                            if t["title"] == s["title"]), None)}
                for s in before["sheets"]
            },
            "formula_errors_src": len(before["formula_errors"]),
            "formula_errors_out": len(after["formula_errors"]),
            "formula_cells_src": before["formula_cells"],
            "formula_cells_out": after["formula_cells"],
        },
    }
    if a.verify_out:
        json.dump(verify, open(a.verify_out, "w", encoding="utf-8"),
                  ensure_ascii=False, indent=2)
    print(json.dumps(verify["checks"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
